"""Combine DCF + comps into a recommendation, charts, and an Excel model."""
from __future__ import annotations

import os
from dataclasses import dataclass

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from .assumptions import Assumptions
from .comps import run_comps, CompsResult
from .data import CompanySnapshot
from .dcf import run_dcf, scenarios, sensitivity, DCFResult

# Blend weights for the headline fair value.
W_DCF, W_COMPS = 0.60, 0.40

# Recommendation bands on upside-to-fair-value.
BUY_ABOVE = 0.15
SELL_BELOW = -0.10


@dataclass
class Report:
    snap: CompanySnapshot
    assumptions: Assumptions
    dcf: DCFResult
    comps: CompsResult
    scenarios: dict
    sensitivity: pd.DataFrame
    fair_value: float
    upside: float
    rating: str


def recommend(upside: float) -> str:
    if upside >= BUY_ABOVE:
        return "BUY"
    if upside <= SELL_BELOW:
        return "SELL"
    return "HOLD"


def build_report(snap: CompanySnapshot, a: Assumptions, peers=None) -> Report:
    dcf = run_dcf(snap, a)
    sc = scenarios(snap, a)
    sens = sensitivity(snap, a)
    try:
        comps = run_comps(snap, peers)
    except Exception as e:  # network / data failure -> DCF-only report
        print(f"[!] comps unavailable ({type(e).__name__}); reporting DCF only.")
        comps = None

    dcf_val = dcf.fair_value_gordon
    comps_val = comps.blended_value if comps is not None else float("nan")
    if comps_val == comps_val:  # not NaN
        fair = W_DCF * dcf_val + W_COMPS * comps_val
    else:
        fair = dcf_val
    upside = fair / snap.price - 1
    return Report(
        snap=snap, assumptions=a, dcf=dcf, comps=comps, scenarios=sc,
        sensitivity=sens, fair_value=fair, upside=upside, rating=recommend(upside),
    )


def print_summary(rep: Report) -> None:
    s, d, c = rep.snap, rep.dcf, rep.comps
    line = "=" * 64
    print(line)
    print(f" {s.name}  ({s.ticker})   valuation summary")
    print(line)
    print(f" Market price            : INR {s.price:,.0f}")
    print(f" Shares outstanding      : {s.shares/1e7:,.1f} crore")
    print(f" Market cap              : INR {s.market_cap:,.0f} cr")
    print(f" Net debt                : INR {s.net_debt:,.0f} cr")
    print(f" Latest revenue / EBITDA : INR {s.revenue:,.0f} cr / INR {s.ebitda:,.0f} cr")
    print("-" * 64)
    print(f" WACC                    : {d.wacc:.2%}   (Ke {rep.assumptions.cost_of_equity():.2%})")
    print(f" Terminal growth         : {rep.assumptions.terminal_growth:.2%}")
    print(f" DCF fair value (Gordon) : INR {d.fair_value_gordon:,.0f}   ({d.upside_gordon:+.1%})")
    print(f" DCF fair value (exit x) : INR {d.fair_value_exit:,.0f}")
    print(f" DCF scenarios bear/base/bull: INR {rep.scenarios['bear']:,.0f} / "
          f"{rep.scenarios['base']:,.0f} / {rep.scenarios['bull']:,.0f}")
    print("-" * 64)
    if c is not None:
        print(f" Peer median EV/EBITDA   : {c.median_ev_ebitda:.1f}x   (Trent {c.trent_ev_ebitda:.1f}x)")
        print(f" Peer median P/E         : {c.median_pe:.1f}x   (Trent {c.trent_pe:.1f}x)")
        print(f" Comps value EV/EBITDA   : INR {c.implied_value_ev_ebitda:,.0f}")
        print(f" Comps value P/E         : INR {c.implied_value_pe:,.0f}")
        print(f" Comps blended           : INR {c.blended_value:,.0f}")
    else:
        print(" Comps                   : unavailable (offline)")
    print("=" * 64)
    print(f" BLENDED FAIR VALUE      : INR {rep.fair_value:,.0f}")
    print(f" Upside / (downside)     : {rep.upside:+.1%}")
    print(f" RECOMMENDATION          : {rep.rating}")
    print("=" * 64)


def football_field(rep: Report, out_dir: str) -> str:
    s, d, c = rep.snap, rep.dcf, rep.comps
    rows = [
        ("DCF (bear-bull)", rep.scenarios["bear"], rep.scenarios["bull"]),
        ("DCF exit-multiple", min(d.fair_value_gordon, d.fair_value_exit), max(d.fair_value_gordon, d.fair_value_exit)),
    ]
    if c is not None:
        rows += [
            ("Comps EV/EBITDA", c.implied_value_ev_ebitda * 0.9, c.implied_value_ev_ebitda * 1.1),
            ("Comps P/E", c.implied_value_pe * 0.9, c.implied_value_pe * 1.1),
        ]
    fig, ax = plt.subplots(figsize=(9, 4.2))
    for i, (label, lo, hi) in enumerate(rows):
        lo, hi = sorted((lo, hi))
        ax.barh(i, hi - lo, left=lo, height=0.5, color="#3b6ea5", alpha=0.85)
        ax.text((lo + hi) / 2, i, f"{lo:,.0f}-{hi:,.0f}", va="center", ha="center", color="white", fontsize=8)
    ax.axvline(s.price, color="crimson", lw=2, label=f"market price {s.price:,.0f}")
    ax.axvline(rep.fair_value, color="green", lw=2, ls="--", label=f"blended fair {rep.fair_value:,.0f}")
    ax.set_yticks(range(len(rows)))
    ax.set_yticklabels([r[0] for r in rows])
    ax.set_xlabel("value per share (INR)")
    ax.set_title(f"{s.name}: valuation football field — {rep.rating}")
    ax.legend(loc="lower right", fontsize=8)
    ax.grid(axis="x", alpha=0.25)
    fig.tight_layout()
    path = os.path.join(out_dir, "football_field.png")
    fig.savefig(path, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return path


def sensitivity_heatmap(rep: Report, out_dir: str) -> str:
    grid = rep.sensitivity.astype(float)
    fig, ax = plt.subplots(figsize=(7, 4.6))
    im = ax.imshow(grid.values, cmap="RdYlGn", aspect="auto")
    ax.set_xticks(range(len(grid.columns)))
    ax.set_xticklabels(grid.columns)
    ax.set_yticks(range(len(grid.index)))
    ax.set_yticklabels(grid.index)
    ax.set_xlabel("terminal growth g")
    ax.set_ylabel("WACC")
    for i in range(grid.shape[0]):
        for j in range(grid.shape[1]):
            ax.text(j, i, f"{grid.values[i, j]:,.0f}", ha="center", va="center", fontsize=8)
    ax.set_title("DCF fair value (INR/share): WACC x terminal growth")
    fig.colorbar(im, ax=ax, shrink=0.8)
    fig.tight_layout()
    path = os.path.join(out_dir, "dcf_sensitivity.png")
    fig.savefig(path, dpi=130, bbox_inches="tight")
    plt.close(fig)
    return path


def export_excel(rep: Report, out_dir: str) -> str:
    """Write a formatted multi-sheet Excel model."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    s, d, c, a = rep.snap, rep.dcf, rep.comps, rep.assumptions
    wb = Workbook()
    hdr = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="3B6EA5")

    def style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = hdr
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    summary_rows = [
        ("Company", s.name), ("Ticker", s.ticker), ("Market price (INR)", round(s.price, 1)),
        ("Shares (cr)", round(s.shares / 1e7, 2)), ("Market cap (INR cr)", round(s.market_cap, 0)),
        ("Net debt (INR cr)", round(s.net_debt, 0)), ("WACC", round(d.wacc, 4)),
        ("Terminal growth", a.terminal_growth),
        ("DCF fair value (Gordon)", round(d.fair_value_gordon, 0)),
        ("DCF fair value (exit x)", round(d.fair_value_exit, 0)),
        ("Comps blended value", round(c.blended_value, 0)),
        ("BLENDED FAIR VALUE", round(rep.fair_value, 0)),
        ("Upside / (downside)", round(rep.upside, 4)),
        ("RECOMMENDATION", rep.rating),
    ]
    ws.append(["Metric", "Value"])
    for r in summary_rows:
        ws.append(list(r))
    style_header(ws, 2)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    # --- DCF projections sheet ---
    ws2 = wb.create_sheet("DCF")
    proj = d.projections.round(1)
    ws2.append(list(proj.columns))
    for _, row in proj.iterrows():
        ws2.append([round(float(x), 2) for x in row.values])
    style_header(ws2, len(proj.columns))

    # --- Comps sheet ---
    if c is not None:
        ws3 = wb.create_sheet("Comps")
        tbl = c.peer_table.copy()
        ws3.append(list(tbl.columns))
        for _, row in tbl.iterrows():
            ws3.append([row[col] if pd.notna(row[col]) else None for col in tbl.columns])
        ws3.append([])
        ws3.append(["MEDIAN EV/EBITDA", round(c.median_ev_ebitda, 2)])
        ws3.append(["MEDIAN P/E", round(c.median_pe, 2)])
        style_header(ws3, len(tbl.columns))

    # --- Sensitivity sheet ---
    ws4 = wb.create_sheet("Sensitivity")
    grid = rep.sensitivity
    ws4.append([grid.index.name] + list(grid.columns))
    for idx, row in grid.iterrows():
        ws4.append([idx] + [round(float(x), 1) for x in row.values])
    style_header(ws4, len(grid.columns) + 1)

    path = os.path.join(out_dir, "Trent_valuation_model.xlsx")
    wb.save(path)
    return path
